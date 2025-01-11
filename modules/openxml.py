import openpyxl
import os

def exel_tabl():    
    directory_file = str(os.getcwd()) + '/xml_files'

    file = os.listdir(directory_file)

    if not len(file) == 0:
        # Сортировать файлы по времени изменения
        file.sort(key=lambda i: os.path.getmtime(os.path.join(directory_file, i)))
        last_file = file[-1]

        work_xml = openpyxl.load_workbook(os.path.join(directory_file, last_file))

        empl = work_xml['Лист 1']
        my_dict = {}
        for row in range(4, empl.max_row+1):
            row_dict = {} # для каждой строки
            # Определяем ключи словаря на основе порядка столбцов в таблице
            for x in range(1, empl.max_column+1):
                if x == 1:
                    key = 'Марка'
                elif x == 2:
                    key = 'Серийный №'
                elif x == 3:
                    key = 'Назначение устройства'
                elif x == 4:
                    key = 'Сетевое имя'
                elif x == 5:
                    key = 'Внутреннее помещение'
                elif x == 6:
                    key = 'Расположен в шкафу'
                elif x == 7:
                    key = 'Инвентарный №'
                elif x == 8:
                    key = 'Учетный номер маркировки'
                else:
                    key = f'Столбец {x}'  # Ключ для остальных столбцов
                row_dict[key] = empl.cell(row=row, column=x).value # Добавляем значение в словарь строки
            my_dict[f'#{row}'] = row_dict
        return my_dict
    return False